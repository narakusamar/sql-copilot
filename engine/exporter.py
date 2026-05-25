import csv
import io
from datetime import datetime


def export_csv(columns: list[str], rows: list[tuple], filename: str = "") -> bytes:
    """导出为 CSV（UTF-8 BOM，Excel 兼容中文）"""
    buf = io.StringIO()
    buf.write("﻿")  # BOM for Excel UTF-8 recognition
    writer = csv.writer(buf)
    if columns:
        writer.writerow(columns)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")


def export_excel(columns: list[str], rows: list[tuple], filename: str = "") -> bytes:
    """导出为 Excel xlsx（需要 openpyxl）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "查询结果"

    # Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4",
                              fill_type="solid")
    for ci, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)

    # Auto-width
    for ci, col in enumerate(columns, 1):
        max_len = len(str(col))
        for row in rows[:100]:
            cell_len = len(str(row[ci - 1])) if row[ci - 1] is not None else 0
            max_len = max(max_len, cell_len)
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(
            max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def default_filename(prefix: str = "export") -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{ts}"
